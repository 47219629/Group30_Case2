from openpyxl import load_workbook

#Import the excel workbook and specify the sheet name
workbook = load_workbook('QUB_DCF2.xlsx', data_only=True)

sheet = workbook['DCF']

#Extract the relevant data
data = []
for row in sheet.iter_rows(min_row=3, max_row=13, min_col=2, max_col=8, values_only=True):
    filtered_row = [cell for i, cell in enumerate(row) if i != 1]
    data.append(filtered_row)

#Use the first row as the header
headers = data[0]
data = data[1:]  

#Round to 0 decimal places
rounded_data = []
for row in data:
    rounded_row = [int(cell) if isinstance(cell, (int, float)) else cell for cell in row]
    rounded_data.append(rounded_row)

#Adding the data to the PDF document
pdf.set_xy(20, 150)
pdf.set_font("Arial", "B", 10)
pdf.cell(0, 10, "Discounted Cash Flow (DCF) Table", ln=True)

#Add table headers
pdf.set_font("Arial", "B", 8)
pdf.set_fill_color(147, 185, 223)
first_col_width = 35
other_col_width = 15
row_height = 5 

#Format headers
pdf.cell(first_col_width, row_height, str(headers[0]) if headers[0] is not None else "", border="T,B", align='L', fill=True)
for header in headers[1:]:
    pdf.cell(other_col_width, row_height, str(header) if header is not None else "", border="T,B", align='C', fill=True)
pdf.ln()

#Add and format table rows
pdf.set_font("Arial", "", 8)
for row_index, row in enumerate(rounded_data):
    if row_index == 4:
        pdf.set_font("Arial", "B", 8)
        pdf.cell(first_col_width, row_height, str(row[0]) if row[0] is not None else "", border="T,B", align='L')
        for cell in row[1:]:
            pdf.cell(other_col_width, row_height, str(cell) if cell is not None else "", border="T,B", align='C')
        pdf.set_font("Arial", "", 8)
    elif row_index == len(rounded_data) - 1: 
        pdf.set_font("Arial", "B", 8) 
        pdf.cell(first_col_width, row_height, str(row[0]) if row[0] is not None else "", border="T,B", align='L')
        for cell in row[1:]:
            pdf.cell(other_col_width, row_height, str(cell) if cell is not None else "", border="T,B", align='C')
        pdf.set_font("Arial", "", 8) 
    else: 
        pdf.cell(first_col_width, row_height, str(row[0]) if row[0] is not None else "", border=0, align='L')
        for cell in row[1:]:
            pdf.cell(other_col_width, row_height, str(cell) if cell is not None else "", border=0, align='C')
    pdf.ln()